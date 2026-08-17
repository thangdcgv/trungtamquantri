from fastapi import APIRouter, Request, Form, UploadFile, File, HTTPException, status, Depends
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from config import supabase
from app.auth import require_login 

import pandas as pd
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
import io
from datetime import datetime

router = APIRouter(prefix="/warranty", tags=["Warranty"])
templates = Jinja2Templates(directory="app/templates")

# --- HÀM BỔ TRỢ CHUẨN HÓA DỮ LIỆU ---
def format_phone_number(phone: str) -> str:
    """Chuẩn hóa SĐT: tự thêm số 0 ở đầu nếu bị mất số 0 do đọc Excel/nhập thiếu."""
    phone_raw = str(phone or "").strip()
    if phone_raw and not phone_raw.startswith('0') and len(phone_raw) in [9, 10]:
        phone_raw = "0" + phone_raw
    return phone_raw

def validate_condition(cond_str: str) -> str:
    """Đảm bảo condition luôn thuộc ['Mới', 'Đã qua sử dụng'] để không vi phạm CHECK constraint của DB."""
    cond_clean = str(cond_str or "").strip()
    if "mới" in cond_clean.lower():
        return "Mới"
    return "Đã qua sử dụng"

def get_active_policies():
    """Lấy danh sách các gói chính sách bảo hành đang hoạt động từ bảng warranty_policy."""
    try:
        res = (
            supabase.table("warranty_policy")
            .select("*")
            .eq("is_active", True)
            .order("created_at", desc=True)
            .execute()
        )
        return res.data if res.data else []
    except Exception as e:
        print(f"❌ Lỗi lấy chính sách: {e}")
        return []

def get_staff_display_name(current_user: dict) -> str:
    """Lấy tên hiển thị của staff linh hoạt theo ho_ten hoặc username."""
    if not current_user:
        return "Hệ thống"
    return current_user.get("ho_ten") or current_user.get("username") or "Hệ thống"

# --- 1. ROUTE XEM DANH SÁCH PHIẾU BẢO HÀNH ---
@router.get("/list")
async def get_warranty_list(
    request: Request,
    q: str = None,
    created_by: str = None,
    start_date: str = None,
    end_date: str = None,
    current_user: dict = Depends(require_login)
):
    try:
        # 1. TRUY VẤN DANH SÁCH NGƯỜI TẠO PHIẾU NỘI BỘ
        creators_res = supabase.table("warranty_records").select("staff_name").execute()

        creators_set = set()
        if creators_res.data:
            for item in creators_res.data:
                val = item.get("staff_name")
                # Chuỗi String chuẩn
                if isinstance(val, str) and val.strip():
                    creators_set.add(val.strip())
                # Dạng Dict (JSON Object từ Supabase)
                elif isinstance(val, dict):
                    name_str = val.get("display_name") or val.get("name") or str(val)
                    if name_str and isinstance(name_str, str) and name_str.strip():
                        creators_set.add(name_str.strip())

        creators_list = sorted(list(creators_set))

        # 2. XÂY DỰNG TRUY VẤN BỘ LỌC KẾT QUẢ
        query = supabase.table("warranty_records").select("*")

        # Lọc theo từ khóa (S/N, Tên KH, SĐT, Model)
        if q and q.strip():
            # Loại bỏ dấu phẩy để tránh vỡ cú pháp Postgrest in .or_()
            clean_q = q.strip().replace(",", " ")
            keyword = f"%{clean_q}%"
            query = query.or_(
                f"serial_number.ilike.{keyword},"
                f"customer_name.ilike.{keyword},"
                f"phone_number.ilike.{keyword},"
                f"model_name.ilike.{keyword}"
            )

        # Lọc theo Người tạo phiếu (Chỉ lọc nếu giá trị không rỗng)
        if created_by and created_by.strip():
            query = query.eq("staff_name", created_by.strip())

        # Lọc theo Thời gian (Từ ngày - Đến ngày)
        if start_date and start_date.strip():
            query = query.gte("purchase_date", start_date.strip())
        if end_date and end_date.strip():
            query = query.lte("purchase_date", end_date.strip())

        # Trả về kết quả mới nhất lên đầu
        records_res = query.order("created_at", desc=True).limit(20).execute()

        return templates.TemplateResponse(
            request=request,
            name="warranty_list.html",
            context={
                "items": records_res.data or [],
                "creators": creators_list,
                "current_user": current_user
            }
        )

    except Exception as e:
        return templates.TemplateResponse(
            "warranty_list.html", 
            {
                "request": request,
                "items": [],
                "creators": [],
                "error_msg": str(e)
            },
            status_code=500
        )

# --- 2. ROUTE HIỂN THỊ FORM TẠO PHIẾU MỚI ---
@router.get("/create", response_class=HTMLResponse)
async def render_create_page(request: Request, current_user: dict = Depends(require_login)):
    policies = get_active_policies()
    return templates.TemplateResponse(
        request=request,
        name="create_warranty.html",
        context={
            "current_user": current_user,
            "policies": policies
        }
    )

# --- 3. ROUTE TẢI FILE EXCEL MẪU ---
@router.get("/api/download-template")
async def download_excel_template():
    try:
        active_policies = get_active_policies()
        policy_names = [p['policy_name'] for p in active_policies if p.get('policy_name')] or [
            "Canon - Tiêu Chuẩn", "Epson - EcoTank", "Laser - Khai Thác"
        ]
        
        categories = ["Máy In Phun", "Máy Laser", "Máy In Kim", "Máy In Tem Nhãn"]
        conditions = ["Mới", "Đã qua sử dụng"]
        columns = [
            "Tên khách hàng", "Số điện thoại", "Ngày mua (YYYY-MM-DD)", 
            "Loại máy", "Tình trạng", "Model máy", "Số Serial (S/N)", 
            "Số trang ban đầu", "Tên Gói Bảo Hành"
        ]
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Template"
        ws.append(columns)
        
        ws_data = wb.create_sheet("HiddenData")
        for i, val in enumerate(categories, 1): ws_data.cell(row=i, column=1, value=val)
        for i, val in enumerate(conditions, 1): ws_data.cell(row=i, column=2, value=val)
        for i, val in enumerate(policy_names, 1): ws_data.cell(row=i, column=3, value=val)
        ws_data.sheet_state = 'hidden'

        dv_cat = DataValidation(type="list", formula1=f"HiddenData!$A$1:$A${len(categories)}", allow_blank=True)
        ws.add_data_validation(dv_cat)
        dv_cat.add("D2:D500")

        dv_cond = DataValidation(type="list", formula1=f"HiddenData!$B$1:$B${len(conditions)}", allow_blank=True)
        ws.add_data_validation(dv_cond)
        dv_cond.add("E2:E500")

        if policy_names:
            dv_pol = DataValidation(type="list", formula1=f"HiddenData!$C$1:$C${len(policy_names)}", allow_blank=True)
            ws.add_data_validation(dv_pol)
            dv_pol.add("I2:I500")

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=mau_bao_hanh_dai_thanh.xlsx"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Lỗi tạo file mẫu: {str(e)}")

# --- 4. ROUTE XỬ LÝ TẠO PHIẾU BẢO HÀNH ---
@router.post("/api/create")
async def create_warranty_record(
    customer_name: str = Form(...),
    phone_number: str = Form(...),
    purchase_date: str = Form(None),
    applied_policy_name: str = Form("NULL"),
    category: str = Form("Máy In Phun"),
    condition: str = Form("Đã qua sử dụng"),
    model_name: str = Form(...),
    serial_number: str = Form(...),
    warranty_months: int = Form(12),
    head_warranty_months: int = Form(0),
    cartridge_warranty_months: int = Form(0),
    initial_counter: int = Form(0),
    page_limit_body: int = Form(10000),
    page_limit_head: int = Form(3000),
    no_page_limit: str = Form("false"),
    current_user: dict = Depends(require_login)
):
    try:
        is_no_limit = str(no_page_limit).lower() == "true"
        clean_sn = serial_number.strip().upper()
        clean_phone = format_phone_number(phone_number)
        clean_customer = customer_name.strip().title()
        clean_model = model_name.strip().upper()
        clean_condition = validate_condition(condition)
        staff_display = get_staff_display_name(current_user)
        
        # Kiểm tra trùng Serial Number
        existing = supabase.table("warranty_records").select("id").eq("serial_number", clean_sn).execute()
        if existing.data:
            return JSONResponse(
                status_code=400,
                content={"success": False, "message": f"🚫 TRÙNG DỮ LIỆU: Số Serial '{clean_sn}' đã tồn tại!"}
            )

        final_head_m = 0 if category == "Máy Laser" else head_warranty_months
        final_cart_m = cartridge_warranty_months if category == "Máy Laser" else 0

        now_str = datetime.now().strftime("%H:%M %d/%m/%Y")
        log_entry = f"• [{now_str}] {staff_display}: KHỞI TẠO (Gói: {applied_policy_name}) | Máy: {clean_model} - S/N: {clean_sn}"

        payload = {
            "customer_name": clean_customer,
            "phone_number": clean_phone,
            "purchase_date": purchase_date if purchase_date else datetime.now().strftime("%Y-%m-%d"),
            "applied_policy_name": applied_policy_name,
            "category": category,
            "condition": clean_condition,
            "model_name": clean_model,
            "serial_number": clean_sn,
            "warranty_months": warranty_months,
            "head_warranty_months": final_head_m,
            "cartridge_warranty_months": final_cart_m,
            "initial_counter": initial_counter,
            "current_counter": initial_counter,
            "page_limit_body": page_limit_body,
            "page_limit_head": page_limit_head,
            "staff_name": staff_display,
            "no_page_limit": is_no_limit,
            "edit_log": log_entry,
            "pages_updated_at": datetime.now().isoformat()
        }

        res = supabase.table("warranty_records").insert(payload).execute()
        
        if res.data:
            return RedirectResponse(
                url="/warranty/create?status=created", 
                status_code=status.HTTP_303_SEE_OTHER
            )
        else:
            raise HTTPException(status_code=500, detail="Không thể tạo dữ liệu bảo hành")

    except Exception as e:
        return JSONResponse(status_code=500, content={"success": False, "message": f"Lỗi hệ thống: {str(e)}"})

# --- 5. ROUTE NHẬP DANH SÁCH TỪ EXCEL ---
@router.post("/api/import-excel")
async def import_warranties_from_excel(
    file: UploadFile = File(...),
    current_user: dict = Depends(require_login)
):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Vui lòng tải lên file định dạng Excel (.xlsx, .xls)")

    try:
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents), dtype={"Số điện thoại": str})
        
        active_policies = get_active_policies()
        records_to_insert = []
        now_str = datetime.now().strftime("%H:%M %d/%m/%Y")
        staff_display = get_staff_display_name(current_user)

        for idx, row in df.iterrows():
            sn_clean = str(row.get("Số Serial (S/N)", "")).strip().upper()
            if not sn_clean or sn_clean == "NAN":
                continue

            policy_name = str(row.get("Tên Gói Bảo Hành", "")).strip()
            p_info = next((p for p in active_policies if p.get('policy_name') == policy_name), None) if active_policies else None

            cat = str(row.get("Loại máy", "Máy In Phun")).strip()
            cond_raw = str(row.get("Tình trạng", "Đã qua sử dụng")).strip()
            cond = validate_condition(cond_raw)
            
            w_months = int(p_info.get('warranty_months', 12)) if p_info else 12
            h_months = int(p_info.get('head_warranty_months', 0)) if p_info and cat != "Máy Laser" else 0
            c_months = int(p_info.get('cartridge_warranty_months', 0)) if p_info and cat == "Máy Laser" else 0
            
            p_body = int(p_info.get('page_limit_body', 10000)) if p_info else 10000
            p_head = int(p_info.get('page_limit_head', 3000)) if p_info else 3000
            no_limit = p_info.get('no_page_limit', False) if p_info else False

            init_cnt = int(row.get("Số trang ban đầu") if pd.notnull(row.get("Số trang ban đầu")) else 0)

            record = {
                "customer_name": str(row.get("Tên khách hàng", "")).strip().title(),
                "phone_number": format_phone_number(row.get("Số điện thoại")),
                "purchase_date": str(row.get("Ngày mua (YYYY-MM-DD)", datetime.now().strftime("%Y-%m-%d"))).split(' ')[0],
                "category": cat,
                "condition": cond,
                "model_name": str(row.get("Model máy", "")).strip().upper(),
                "serial_number": sn_clean,
                "initial_counter": init_cnt,
                "current_counter": init_cnt,
                "warranty_months": w_months,
                "head_warranty_months": h_months,
                "cartridge_warranty_months": c_months,
                "page_limit_body": p_body,
                "page_limit_head": p_head,
                "no_page_limit": no_limit,
                "applied_policy_name": policy_name if policy_name else "Nhập từ Excel",
                "staff_name": staff_display,
                "edit_log": f"• [{now_str}] {staff_display}: IMPORT EXCEL",
                "pages_updated_at": datetime.now().isoformat()
            }
            records_to_insert.append(record)

        if not records_to_insert:
            raise HTTPException(status_code=400, detail="Không có dữ liệu hợp lệ trong file Excel")

        res = supabase.table("warranty_records").upsert(records_to_insert, on_conflict="serial_number").execute()
        return {"success": True, "imported_count": len(res.data) if res.data else len(records_to_insert)}

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Lỗi đọc file Excel: {str(e)}")

# --- 6. ROUTE RENDER CHI TIẾT PHIẾU BẢO HÀNH ---
@router.get("/{warranty_id}", response_class=HTMLResponse)
async def render_warranty_detail(
    request: Request, 
    warranty_id: str,
    current_user: dict = Depends(require_login)
):
    res = supabase.table("warranty_records").select("*").eq("id", warranty_id).single().execute()
    if not res.data:
        raise HTTPException(status_code=404, detail="Không tìm thấy phiếu bảo hành")
        
    policies = get_active_policies()
    return templates.TemplateResponse(
        request=request,
        name="detail_warranty.html",
        context={
            "item": res.data,
            "policies": policies,
            "current_user": current_user
        }
    )

# --- 7. ROUTE CẬP NHẬT PHIẾU BẢO HÀNH ---
@router.post("/api/update/{warranty_id}")
async def update_warranty_record(
    warranty_id: str,
    customer_name: str = Form(...),
    phone_number: str = Form(...),
    purchase_date: str = Form(...),
    applied_policy_name: str = Form(...),
    category: str = Form(...),
    condition: str = Form(...),
    model_name: str = Form(...),
    warranty_months: int = Form(12),
    head_warranty_months: int = Form(0),
    cartridge_warranty_months: int = Form(0),
    initial_counter: int = Form(0),
    current_counter: int = Form(0),
    page_limit_body: int = Form(10000),
    page_limit_head: int = Form(3000),
    no_page_limit: str = Form("false"),
    current_user: dict = Depends(require_login)
):
    try:
        old_res = supabase.table("warranty_records").select("*").eq("id", warranty_id).single().execute()
        if not old_res.data:
            raise HTTPException(status_code=404, detail="Không tìm thấy phiếu bảo hành để cập nhật")

        old_data = old_res.data
        staff_display = get_staff_display_name(current_user)

        # -------------------------------------------------------------
        # 🔒 KIỂM TRA QUYỀN CHỈNH SỬA
        # -------------------------------------------------------------
        user_role = current_user.get("role", "")
        is_admin = user_role in ["Admin", "Super Admin", "System Admin"]
        
        # So sánh tên nhân viên tạo phiếu với người đang đăng nhập
        creator_name = old_data.get("staff_name")
        is_owner = creator_name and (creator_name.strip().lower() == staff_display.strip().lower())

        if not (is_admin or is_owner):
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN, 
                detail="Bạn không có quyền chỉnh sửa phiếu bảo hành của nhân viên khác!"
            )
        # -------------------------------------------------------------

        new_customer = customer_name.strip().title()
        new_phone = format_phone_number(phone_number)
        new_model = model_name.strip().upper()
        new_condition = validate_condition(condition)
        is_no_limit = str(no_page_limit).lower() == "true"

        changes = []
        fields_map = {
            "customer_name": ("Tên KH", new_customer),
            "phone_number": ("SĐT", new_phone),
            "model_name": ("Model", new_model),
            "applied_policy_name": ("Gói BH", applied_policy_name),
            "condition": ("Tình trạng", new_condition),
            "warranty_months": ("BH Cơ", warranty_months),
            "head_warranty_months": ("BH Đầu/Mực", head_warranty_months),
            "initial_counter": ("Số trang ban đầu", initial_counter),
            "current_counter": ("Số trang hiện tại", current_counter),
        }
        for k, (label, new_val) in fields_map.items():
            old_val = old_data.get(k)
            if str(old_val).strip() != str(new_val).strip():
                changes.append(f"{label}: {old_val} ➔ {new_val}")

        now_str = datetime.now().strftime("%H:%M %d/%m/%Y")
        detail_str = f" | Thay đổi: {', '.join(changes)}" if changes else " | Không thay đổi nội dung"
        log_entry = f"• [{now_str}] {staff_display}: CẬP NHẬT (Gói: {applied_policy_name}){detail_str}"
        
        existing_log = str(old_data.get("edit_log", "")).strip()
        full_log = f"{log_entry}\n{existing_log}" if existing_log and existing_log.lower() != "none" else log_entry

        final_head_m = 0 if category == "Máy Laser" else head_warranty_months
        final_cart_m = cartridge_warranty_months if category == "Máy Laser" else 0

        update_payload = {
            "customer_name": new_customer,
            "phone_number": new_phone,
            "purchase_date": purchase_date,
            "applied_policy_name": applied_policy_name,
            "category": category,
            "condition": new_condition,
            "model_name": new_model,
            "warranty_months": warranty_months,
            "head_warranty_months": final_head_m,
            "cartridge_warranty_months": final_cart_m,
            "initial_counter": initial_counter,
            "current_counter": current_counter,
            "page_limit_body": page_limit_body,
            "page_limit_head": page_limit_head,
            "no_page_limit": is_no_limit,
            "staff_name": staff_display,
            "edit_log": full_log
        }

        if old_data.get("current_counter") != current_counter:
            update_payload["pages_updated_at"] = datetime.now().isoformat()

        supabase.table("warranty_records").update(update_payload).eq("id", warranty_id).execute()
        
        return RedirectResponse(
            url=f"/warranty/{warranty_id}?status=updated", 
            status_code=status.HTTP_303_SEE_OTHER
        )

    except HTTPException:
        raise
    except Exception as e:
        return JSONResponse(status_code=500, content={"success": False, "message": str(e)})


# --- 8. ROUTE API XÓA PHIẾU BẢO HÀNH ---
@router.delete("/api/delete/{warranty_id}")
async def delete_warranty_record(
    warranty_id: str,
    current_user: dict = Depends(require_login)
):
    try:
        # Fetch dữ liệu cũ để lấy thông tin người tạo trước khi thực hiện xóa
        old_res = supabase.table("warranty_records").select("*").eq("id", warranty_id).single().execute()
        if not old_res.data:
            return JSONResponse(status_code=404, content={"success": False, "message": "Không tìm thấy phiếu bảo hành để xóa"})

        old_data = old_res.data
        staff_display = get_staff_display_name(current_user)

        # -------------------------------------------------------------
        # 🔒 KIỂM TRA QUYỀN XÓA
        # -------------------------------------------------------------
        user_role = current_user.get("role", "")
        is_admin = user_role in ["Admin", "Super Admin", "System Admin"]
        
        creator_name = old_data.get("staff_name")
        is_owner = creator_name and (creator_name.strip().lower() == staff_display.strip().lower())

        if not (is_admin or is_owner):
            return JSONResponse(
                status_code=403, 
                content={"success": False, "message": "Bạn không có quyền xóa phiếu bảo hành của người khác!"}
            )
        # -------------------------------------------------------------

        res = supabase.table("warranty_records").delete().eq("id", warranty_id).execute()
        if res.data:
            return {"success": True, "message": "Xóa phiếu bảo hành thành công"}
            
        return JSONResponse(status_code=400, content={"success": False, "message": "Xóa phiếu không thành công"})

    except Exception as e:
        return JSONResponse(status_code=500, content={"success": False, "message": str(e)})